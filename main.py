"""
resx_helper -

Author:JASON
Date:2024/11/04
"""

import os
import json
import re
import openpyxl
import pandas as pd

"""
    读取js文件，返回字典
"""


def read_js_files(folder_path):
    resxdata = []
    js_objects = {}
    sheel_name = ""
    lang_tag = ""
    # 遍历文件夹及子文件夹
    for dirpath, _, filenames in os.walk(folder_path):
        # print(dirpath)
        FolderName = dirpath.split('\\')[-1]
        if FolderName == "resx":
            sheel_name = "global"
            lang_tag = ""
        else:
            lang_tag = FolderName

        for filename in filenames:
            if filename.endswith('.js'):
                file_path = os.path.join(dirpath, filename)
                if FolderName == "resx":
                    # 使用 split 方法按点号分割字符串
                    parts = filename.split('.')
                    if len(parts) == 3:
                        # 取第二个部分（索引为1）
                        lang_tag = parts[1]
                else:
                    # 取文件夹名称不要后缀
                    sheel_name = os.path.splitext(filename)[0]

                # print("sheel_name:" + sheel_name)
                # print("lang_tag:" + lang_tag)
                # 读取文件内容
                with open(file_path, 'r', encoding='utf-8') as file:
                    js_content = file.read()
                    js_content = re.sub(r'\\.', '', js_content)
                    match = re.search(r'\{.*?\}', js_content, re.DOTALL)
                    # 存储解析后的字典
                    data_dicts = []

                    if match:
                        # 提取匹配到的JSON字符串
                        json_str = match.group(0)

                        # 将JSON字符串转换为Python字典
                        data_dict = json.loads(json_str)
                        # 判断append_data是否已包含当前sheet_name
                        if sheel_name not in js_objects:
                            js_objects[sheel_name] = [{"lang_tag": lang_tag, "data": data_dict}]
                        else:
                            js_objects[sheel_name].append({"lang_tag": lang_tag, "data": data_dict})

                    else:
                        print("No JSON object found.")

    return js_objects


"""
    将数据转换为表格写入格式
    如：
    [{'lang_tag': 'en-us', 'data': {'greeting': 'Hello', 'welcome': 'Welcome to our website!'}}, {'lang_tag': 'zh-cn', 'data': {'greeting': '你好', 'welcome': '欢迎来到我们的网站！'}}]
    转换成
    [
        ['k', 'en-us', 'zh-cn'],
        ['greeting', 'Hello', 'Welcome to our website!'],
        ['welcome', '你好', 'welcome': '欢迎来到我们的网站！']
    ]
"""


def convert_data_to_table(data):
    _lang_tag = [item['lang_tag'] for item in data];
    _lang_tag.insert(0, 'k')

    table = [_lang_tag]
    _lt_idx = 1;
    for item in data:
        lang_tag = item['lang_tag']
        data_dict = item['data']
        for key, value in data_dict.items():
            if key not in [row[0] for row in table]:
                _filds = [key]
                for _lt in _lang_tag[1:]:
                    _filds.append('')
                table.append(_filds)
            row = [row for row in table if row[0] == key][0]
            row[_lt_idx] = value
        _lt_idx += 1

    print("==table==", end=" ")
    print(table)
    return table


# 将js文件转换为excel
def convert_js_to_excel():
    current_dir = os.path.dirname(__file__)
    resx_folder_path = os.path.join(current_dir, 'resx')
    js_objects = read_js_files(resx_folder_path)
    # print(js_objects)

    # 判断resx文件是否有resx.xlsx文件，如果有则删除
    excel_file_path = os.path.join(current_dir, 'resx/resx.xlsx')
    if os.path.exists(excel_file_path):
        os.remove(excel_file_path)

    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()

    for sheet_name, data in js_objects.items():
        print("==sheet_name==", end=" ")
        print(sheet_name)
        print("==data==", end=" ")
        print(data)
        # 创建数据
        data_resx = convert_data_to_table(data)

        # 创建第一个工作表并写入数据
        sheet_resx = workbook.create_sheet(sheet_name)
        for row in data_resx:
            sheet_resx.append(row)

        # 保存工作簿
        workbook.save('resx/resx.xlsx')

    # 删除默认的空工作表
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)
    # 保存工作簿
    workbook.save('resx/resx.xlsx')


# 将js内容写入到对应路径文件
def write_file(write_path, js_content):
    # 文件目录不存在则创建
    if not os.path.exists(os.path.dirname(write_path)):
        os.makedirs(os.path.dirname(write_path))
    with open(write_path, 'w', encoding='utf-8') as file:
        file.write(js_content)


# 将excel转换为js文件
def convert_excel_to_js():
    current_dir = os.path.dirname(__file__)
    excel_file_path = os.path.join(current_dir, 'resx/resx.xlsx')
    # pandas 读取excel
    # 加载整个工作簿
    workbook = pd.ExcelFile(excel_file_path)

    sheet_names = workbook.sheet_names
    # 查看工作簿中的所有工作表名称

    for sheet_name in sheet_names:
        # 选择要读取的工作表
        df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        # 获取列名
        lang_tag = df.columns[1:]
        # 读取指定列的数据
        column_data = df['k']

        kValus = column_data.tolist()

        write_path = ""
        jsObjectName = ""
        # 读取第一列的数据
        for _lang in lang_tag:
            if sheet_name == "global":
                write_path = f"resx/{sheet_name}.{_lang}.js"
                jsObjectName = "resx"
            else:
                write_path = f"resx/{_lang}/{sheet_name}.js"
                jsObjectName = "resx_" + sheet_name

            js_content = "var " + jsObjectName + " = {\n"
            for i in range(len(kValus)):
                _resx_vle = df[_lang][i]
                js_content += f"    \"{kValus[i]}\":\"{_resx_vle}\",\n"
            # 去除最后的逗号
            js_content = js_content[:-2]
            js_content += "\n}"
            write_file(write_path, js_content)

        # for row in sheet.iter_rows(min_row=2, values_only=True):
        #     print(row)


chooseAction = input("请输入操作：\n1. js->excel \n2. excel->js\n")
if chooseAction == "1":
    print("js->excel")
    convert_js_to_excel()
    print("处理完成")
else:
    print("excel->js")
    convert_excel_to_js()
    print("处理完成")
